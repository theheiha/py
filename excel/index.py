import xlrd
import datetime
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from threading import Thread
from queue import Queue

class CreateNewExcel(object):
    def __init__(self):
        # 默认表头配置，可根据实际情况修改												
        self.header_titles = ['送货日期','客户简称','生产单号','客户订单号','客户货号','名称','产品规格','长度(mm)','宽度(mm)','高度(mm)','发货数量','计量单位','产品单价','产品总价']
        # 原文件名
        self.file_path = '对账单模板.xls'
        # 默认计量单位，空值自动填充
        self.default_unit = '件'
        self.q = Queue()

    def get_curr_path(self):
        """
        获取当前文件路径
        """
        curr_dir = os.path.dirname(os.path.abspath(__file__))
        result_dir = os.path.join(curr_dir, self.file_path)

        return result_dir

    def get_curr_date(self):
        """
        获取当前日期
        """
        date = datetime.datetime.now()

        return date.strftime('%Y年%m月')
    
    def handle_data(self):
        """
        处理数据
        """
        # 打开excel文件
        workbook = xlrd.open_workbook(self.get_curr_path())
        # 获取excel文件的第一个sheet
        sheet = workbook.sheet_by_index(0)
        # 获取表头行
        header_row = sheet.row_values(0)
        excel_data = {}

        for column in range(sheet.ncols):
            excel_data[f'{header_row[column]}'] = sheet.col_values(column)[1:]
        
        data = {}
        for header_title in self.header_titles:
            if header_title in excel_data:
                data[header_title] = excel_data[header_title]
            else:
                print(f'未找到表头: {header_title}')
                # 未找到表头，填充空值占位
                data[header_title] = [''] * len(sheet.col_values(0)-1)
        
        self.set_specs(data)
        df = pd.DataFrame(data)
        self.set_empty(df)

        # 根据客户简称分组
        grouped = df.groupby('客户简称')
        for name, group in grouped:
            new_data = group.to_dict(orient='list')
            new_title = f'{name}{self.get_curr_date()}对账单'
            print(f'开始处理: {name}')
            self.q.put({'title': new_title, 'data': new_data})
        
    def set_empty(self, df):
        """
        填充空值
        """
        for i in df['计量单位']:
            if i != self.default_unit:
                df['计量单位'] = df['计量单位'].replace(i, self.default_unit)

    def set_specs(self, data):
        """
        设置产品规格
        """
        lengths = data.get('长度(mm)', [])
        widths = data.get('宽度(mm)', [])
        heights = data.get('高度(mm)', [])
        specs = []
        for length, width, height in zip(lengths, widths, heights):
            if length and width and height:
                spec = f'{length}*{width}*{height}'
            elif length and width:
                spec = f'{length}*{width}'
            elif length and height:
                spec = f'{length}*{height}'
            elif width and height:
                spec = f'{width}*{height}'
            elif length:
                spec = f'{length}'
            elif width:
                spec = f'{width}'
            elif height:
                spec = f'{height}'
            else:
                spec = ''
            specs.append(spec)

        data['产品规格'] = specs

    def save_to_excel(self):
        """
        保存到excel
        """
        while True:
            data = self.q.get()
            title = data.get('title', '')
            d = data.get('data', {})

            df = pd.DataFrame(d)
            df['产品总价'] = pd.to_numeric(df['产品总价'])
            new_excel_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), f'{title}.xlsx')
            df.to_excel(new_excel_path, index=False)

            # 打开新的excel
            wb = load_workbook(new_excel_path)
            ws = wb.active

            self.insert_title(ws, title)
            self.insert_total_price(ws, df)

            wb.save(new_excel_path)
            print(f'\n处理完成: {title}')
            self.q.task_done()

    def insert_title(self, ws, title):
        """
        插入标题
        """
        ws.insert_rows(1)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(self.header_titles))
        center_alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=1, column=1, value=title).alignment = center_alignment

    def insert_total_price(self, ws, df):
        """
        插入总金额
        确保表格最后一列为产品总价列
        """
        index = len(df) + 3
        header_titles_len = len(self.header_titles)
        ws.insert_rows(index)
        ws.cell(row=index, column=header_titles_len-1, value='合计')
        start_cell = f'{chr(64+header_titles_len)}3'
        end_cell = f'{chr(64+header_titles_len)}{len(df)+2}'
        ws.cell(row=index, column=header_titles_len, value=f'=SUM({start_cell}:{end_cell})')

    def run(self):
        """
        执行
        """
        t1 = Thread(target=self.handle_data)
        t2 = Thread(target=self.save_to_excel)
        t1.start()
        t2.setDaemon(True)
        t2.start()
        t1.join()
        self.q.join()

if __name__ == '__main__':
    CreateNewExcel().run()
